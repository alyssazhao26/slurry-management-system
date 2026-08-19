r"""Create one self-contained GNEM static dashboard from MySQL or an Excel export.

Examples:
  .venv\Scripts\python static_dashboard\generate_static_dashboard.py --date 2026-08-18
  .venv\Scripts\python static_dashboard\generate_static_dashboard.py --source excel --excel C:\reports\daily.xlsx
"""
import argparse
import html
import os
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

PROJECT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT))

from dotenv import load_dotenv
import mysql.connector

load_dotenv(PROJECT / ".env", override=True)


def as_text(value, fallback="—"):
    if value in (None, ""):
        return fallback
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def as_number(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def read_mysql(day):
    config = {
        "host": os.getenv("DB_HOST", "127.0.0.1"), "port": int(os.getenv("DB_PORT", "3306")),
        "database": os.getenv("DB_NAME", "slurry_management"), "user": os.getenv("DB_USER", "slurry_app"),
        "password": os.getenv("DB_PASSWORD", ""),
    }
    connection = mysql.connector.connect(**config)
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT record_date, shift_name, machine_code, formula_code, batch_number,
                   planned_quantity, actual_quantity, qualified_quantity, qualified_pending,
                   achievement_rate, qualified_rate
            FROM system_import.production_records WHERE record_date = %s
            ORDER BY machine_code, shift_name, batch_number
        """, (day,))
        production = cursor.fetchall()
        cursor.execute("""
            SELECT event_date, start_time, end_time, machine_code, machine_type, event_type,
                   severity, description, is_resolved, responsible_person, target_finish_date,
                   actual_finish_date, solution_provided
            FROM system_import.abnormality_reports WHERE event_date = %s
            ORDER BY start_time, machine_code
        """, (day,))
        events = cursor.fetchall()
        cursor.execute("""
            SELECT event_date, machine_code, event_type, severity, responsible_person, target_finish_date
            FROM system_import.abnormality_reports
            WHERE is_resolved = 'no' ORDER BY event_date DESC, start_time DESC LIMIT 12
        """)
        ongoing = cursor.fetchall()
        return production, events, ongoing
    finally:
        connection.close()


def read_excel(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel input needs openpyxl. Run .venv\\Scripts\\pip install -r requirements.txt first.") from exc
    workbook = load_workbook(path, data_only=True, read_only=True)

    def rows_for(names):
        sheet = next((workbook[name] for name in names if name in workbook.sheetnames), None)
        if not sheet:
            return []
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip().lower() if value is not None else "" for value in next(rows)]
        return [dict(zip(headers, values)) for values in rows if any(value is not None for value in values)]

    production = rows_for(["production_records", "production", "Production"])
    events = rows_for(["abnormality_reports", "events", "Event tracker"])
    ongoing = [row for row in events if str(row.get("is_resolved", "no")).lower() != "yes"]
    return production, events, ongoing


def build_html(day, production, events, ongoing):
    planned = sum(as_number(row.get("planned_quantity")) for row in production)
    actual = sum(as_number(row.get("actual_quantity")) for row in production)
    qualified = sum(as_number(row.get("qualified_quantity")) for row in production)
    pending = sum(1 for row in production if row.get("qualified_pending") in (1, True, "1", "yes", "Yes"))
    achievement = actual / planned if planned else None
    qualified_rate = qualified / actual if actual else None
    open_events = sum(1 for row in events if str(row.get("is_resolved", "no")).lower() != "yes")
    machines = defaultdict(lambda: {"planned": 0, "actual": 0, "events": 0})
    for row in production:
        item = machines[as_text(row.get("machine_code"), "NA")]
        item["planned"] += as_number(row.get("planned_quantity")); item["actual"] += as_number(row.get("actual_quantity"))
    for row in events:
        machines[as_text(row.get("machine_code"), "NA")]["events"] += 1

    def rate(value): return "Pending / 待补录" if value is None else f"{value * 100:.1f}%"
    def table(rows, columns, empty):
        if not rows: return f'<p class="empty">{empty}</p>'
        head = "".join(f"<th>{html.escape(label)}</th>" for _, label in columns)
        body = "".join("<tr>" + "".join(f"<td>{html.escape(as_text(row.get(key)))}</td>" for key, _ in columns) + "</tr>" for row in rows)
        return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"

    machine_cards = "".join(
        f'<article class="machine"><h3>{html.escape(name)}</h3><p>Actual / 实际: <b>{values["actual"]:.2f}</b></p><p>Plan / 计划: {values["planned"]:.2f}</p><p class="{ "warn" if values["events"] else ""}">Events / 事件: {values["events"]}</p></article>'
        for name, values in sorted(machines.items())
    ) or '<p class="empty">No production or event data / 暂无生产或事件数据</p>'
    event_rows = []
    for row in events:
        copied = dict(row)
        copied["event_time"] = f'{as_text(row.get("start_time"))} – {as_text(row.get("end_time"))}'
        copied["status"] = "Resolved / 已解决" if str(row.get("is_resolved", "no")).lower() == "yes" else "Open / 进行中"
        event_rows.append(copied)

    return f'''<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>GNEM Daily Dashboard — {html.escape(str(day))}</title><style>
*{{box-sizing:border-box}} body{{margin:0;background:#06233f;color:#fff;font:16px/1.4 Segoe UI,Arial,sans-serif}} main{{max-width:1600px;margin:auto;padding:30px}} header{{display:flex;justify-content:space-between;align-items:end;gap:20px}} .eyebrow{{color:#74c7ef;font-weight:800;letter-spacing:1px;font-size:13px}} h1{{font-size:42px;margin:3px 0}} .muted{{color:#c9e1ee}} .metrics{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin:22px 0}} .metrics article,.section{{border:1px solid #ffffff42;background:#ffffff12;border-radius:16px;padding:18px}} .metrics b{{display:block;font-size:42px;margin:5px 0}} h2{{margin:0 0 14px;font-size:26px}} .machines{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}} .machine{{background:#fff;color:#102b42;border-radius:12px;padding:14px}} .machine h3{{margin:0;color:#06233f;font-size:24px}} .machine p{{margin:5px 0;font-size:14px}} .warn{{color:#875500;font-weight:800}} .grid{{display:grid;grid-template-columns:3fr 2fr;gap:14px;margin-top:14px}} table{{width:100%;border-collapse:collapse;background:#fff;color:#102b42;font-size:13px}} th,td{{padding:9px;text-align:left;border-bottom:1px solid #dae7ef}} th{{background:#e8f2f8;color:#06233f}} .ongoing{{margin:0;padding:0;list-style:none;display:grid;gap:9px}} .ongoing li{{background:#ffe496;color:#102b42;border-radius:10px;padding:11px}} .empty{{color:#c9e1ee}} footer{{text-align:center;color:#c9e1ee;margin:18px}} @media(max-width:900px){{.metrics,.machines{{grid-template-columns:repeat(2,1fr)}}.grid{{grid-template-columns:1fr}}}} @media print{{body{{background:#fff;color:#102b42}}.metrics article,.section{{border-color:#a9bfd0;background:#fff}}.machine{{border:1px solid #cbdce7}}}}
</style></head><body><main><header><div><div class="eyebrow">GNEM SLURRY PRODUCTION TRACKER / GNEM 浆料生产追踪系统</div><h1>Daily management dashboard / 每日管理看板</h1></div><div><b>{html.escape(str(day))} / 当日</b><div class="muted">Static report / 静态报告</div></div></header>
<section class="metrics"><article><span>Achievement rate / 达成率</span><b>{rate(achievement)}</b><small>Actual ÷ plan / 实际 ÷ 计划</small></article><article><span>Qualified rate / 合格率</span><b>{rate(qualified_rate)}</b><small>{pending} pending / 待补录</small></article><article><span>Open events / 进行中事件</span><b>{open_events}</b><small>Requires follow-up / 需要跟进</small></article><article><span>Production records / 生产记录</span><b>{len(production)}</b><small>Selected day / 所选日期</small></article></section>
<section class="section"><h2>Machine overview / 设备概览</h2><div class="machines">{machine_cards}</div></section>
<div class="grid"><section class="section"><h2>Daily event details / 每日事件详情</h2>{table(event_rows, [("machine_code","Machine / 设备"),("event_time","Time / 时间"),("event_type","Event type / 事件类型"),("severity","Severity / Priority / 严重程度 / 优先级"),("responsible_person","Responsible person / 责任人"),("status","Status / 状态")], "No events / 暂无事件")}</section><section class="section"><h2>Ongoing events / 进行中事件</h2><ul class="ongoing">{''.join(f'<li><b>{html.escape(as_text(row.get("machine_code")))} · {html.escape(as_text(row.get("event_type")))}</b><br><small>Severity / Priority / 严重程度 / 优先级: {html.escape(as_text(row.get("severity")))}<br>Responsible person / 责任人: {html.escape(as_text(row.get("responsible_person")))}<br>Expected finish / 预计完成: {html.escape(as_text(row.get("target_finish_date")))}</small></li>' for row in ongoing) or '<li>No ongoing events / 暂无进行中事件</li>'}</ul></section></div><footer>Generated from controlled factory data. Read-only static dashboard / 从受控工厂数据生成，只读静态看板。</footer></main></body></html>'''


def main():
    parser = argparse.ArgumentParser(description="Generate one GNEM static HTML dashboard.")
    parser.add_argument("--source", choices=("mysql", "excel"), default="mysql")
    parser.add_argument("--date", default=date.today().isoformat(), help="Report date: YYYY-MM-DD")
    parser.add_argument("--excel", help="Excel workbook with production_records and abnormality_reports sheets")
    parser.add_argument("--output", help="Output HTML path")
    args = parser.parse_args()
    if args.source == "excel":
        if not args.excel: parser.error("--excel is required with --source excel")
        production, events, ongoing = read_excel(Path(args.excel))
    else:
        production, events, ongoing = read_mysql(args.date)
    output = Path(args.output) if args.output else PROJECT / "static_dashboard" / "output" / f"GNEM_Daily_Dashboard_{args.date}.html"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(build_html(args.date, production, events, ongoing), encoding="utf-8")
    print(f"Created: {output}")


if __name__ == "__main__":
    main()
